from __future__ import annotations

import math
import re
from copy import copy
from typing import Iterable, Set

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["format_worksheet_as_table"]

_TABLE_STYLE_NAME = "TableStyleMedium16"
_TABLE_NAME_SANITIZER = re.compile(r"[^A-Za-z0-9_]")
_COLUMN_PADDING = 1
_DEFAULT_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 255
_MIN_CHAR_WIDTH = 4
_MAX_HEADER_LINES = 4
_HEADER_FONT_COLOR = "00FFFFFF"
_HEADER_FILL_COLOR = "FF1F3864"
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrapText=True)


def format_worksheet_as_table(
    worksheet: Worksheet,
    *,
    table_name: str | None = None,
) -> None:
    """
    Format a worksheet as a styled Excel table with frozen header and auto-fit columns.

    The function assumes row 1 contains headers and applies:
      * Frozen panes after the header row.
      * Header text wrapping, centering, white font color, and a dark blue background.
      * Column widths sized to the longest data value (excluding the header row),
        widened only as needed to avoid breaking header words and to cap header wrapping
        at 4 lines (with a small minimum width for empty columns).
      * A TableStyleMedium16 table for Excel-native filtering/sorting.
    """
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.freeze_panes = "A2"
    _wrap_header_row(worksheet)
    _autofit_columns(worksheet)
    _apply_table_style(worksheet, requested_name=table_name)


def _wrap_header_row(worksheet: Worksheet) -> None:
    header_rows = worksheet.iter_rows(min_row=1, max_row=1)
    try:
        header = next(header_rows)
    except StopIteration:
        return

    for cell in header:
        alignment = copy(cell.alignment) if cell.alignment else Alignment()
        alignment.horizontal = _HEADER_ALIGNMENT.horizontal
        alignment.vertical = _HEADER_ALIGNMENT.vertical
        alignment.wrapText = _HEADER_ALIGNMENT.wrapText
        cell.alignment = alignment

        font = copy(cell.font) if cell.font else Font()
        font.color = _HEADER_FONT_COLOR
        cell.font = font
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=_HEADER_FILL_COLOR,
            bgColor=_HEADER_FILL_COLOR,
        )


def _autofit_columns(worksheet: Worksheet) -> None:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_col == 0:
        return

    for column_cells in worksheet.iter_cols(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        header_cell = column_cells[0]
        data_cells = column_cells[1:]

        data_width = max(
            (_cell_display_length(cell.value) for cell in data_cells),
            default=0,
        )
        header_text = _cell_text(header_cell.value)
        header_total_length = _cell_display_length(header_text)
        header_longest_word = _longest_word_length(header_text)

        adjusted_width = data_width + _COLUMN_PADDING if data_width else 0
        adjusted_width = max(adjusted_width, header_longest_word)

        if header_total_length and adjusted_width:
            estimated_lines = math.ceil(header_total_length / adjusted_width)
            if estimated_lines > _MAX_HEADER_LINES:
                widened_for_header = math.ceil(
                    header_total_length / _MAX_HEADER_LINES
                )
                adjusted_width = max(adjusted_width, widened_for_header)

        if adjusted_width == 0:
            adjusted_width = _DEFAULT_COLUMN_WIDTH

        adjusted_width = max(adjusted_width, _MIN_CHAR_WIDTH)
        adjusted_width = min(adjusted_width, _MAX_COLUMN_WIDTH)
        column_letter = get_column_letter(header_cell.column)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _apply_table_style(worksheet: Worksheet, requested_name: str | None) -> None:
    min_row = worksheet.min_row or 1
    max_row = worksheet.max_row or 1
    min_col = worksheet.min_column or 1
    max_col = worksheet.max_column or 1
    if max_row < min_row or max_col < min_col:
        return

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    table_range = f"{start}:{end}"
    table_name = _generate_table_name(worksheet, requested_name)

    table = Table(displayName=table_name, ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name=_TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _cell_display_length(value: object) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bytes):
        text = value.decode("utf-8", errors="replace")
    else:
        text = str(value)
    if not text:
        return 0
    return max(len(part) for part in text.splitlines()) or 0


def _cell_text(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _longest_word_length(text: str) -> int:
    return max((len(word) for word in re.split(r"\\s+", text) if word), default=0)


def _generate_table_name(worksheet: Worksheet, requested_name: str | None) -> str:
    base_name = requested_name or f"{worksheet.title}_table"
    base_name = _TABLE_NAME_SANITIZER.sub("_", base_name).strip("_") or "Table"
    if not base_name[0].isalpha() and base_name[0] != "_":
        base_name = f"_{base_name}"
    base_name = base_name[:250]

    existing_names = _collect_table_names(worksheet)
    candidate = base_name
    suffix = 1
    while candidate in existing_names:
        candidate = f"{base_name}_{suffix}"
        suffix += 1
    return candidate


def _collect_table_names(worksheet: Worksheet) -> Set[str]:
    workbook = getattr(worksheet, "parent", None)
    tables: Set[str] = set()
    worksheets: Iterable[Worksheet]
    if workbook is not None:
        worksheets = getattr(workbook, "worksheets", []) or []
    else:
        worksheets = [worksheet]

    for ws in worksheets:
        ws_tables = getattr(ws, "tables", None)
        if not ws_tables:
            continue
        if isinstance(ws_tables, dict):
            tables.update(ws_tables.keys())
        else:
            tables.update(getattr(tbl, "displayName", "") for tbl in ws_tables)

    return {name for name in tables if name}
